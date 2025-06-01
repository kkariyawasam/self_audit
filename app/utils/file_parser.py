import fitz  # PyMuPDF

from PIL import Image
from io import BytesIO
from docx import Document
from openpyxl import load_workbook
import os

def extract_text_from_file(file_path: str, content_type: str) -> str:
    import pytesseract
    ext = file_path.split(".")[-1].lower()

    if ext == "pdf":
        return extract_pdf_text(file_path)
    elif ext == "docx":
        return extract_docx_text(file_path)
    elif ext == "xlsx":
        return extract_xlsx_text(file_path)
    elif ext in ["png", "jpg", "jpeg"]:
        return extract_image_text(file_path)
    else:
        return ""

def extract_pdf_text(file_path: str) -> str:
    text = ""
    doc = fitz.open(file_path)
    for page in doc:
        text += page.get_text()
    return text

def extract_docx_text(file_path: str) -> str:
    doc = Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def extract_xlsx_text(file_path: str) -> str:
    wb = load_workbook(file_path)
    text = ""
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            text += " ".join([str(cell.value) if cell.value else "" for cell in row]) + "\n"
    return text

def extract_image_text(file_path: str) -> str:
    img = Image.open(file_path)
    return pytesseract.image_to_string(img)
